import json
import re
import base64
from datetime import datetime
from uuid import uuid1
from io import BytesIO
from tempfile import NamedTemporaryFile

import openpyxl
from openpyxl import load_workbook

from app.domain.models import User, Vacation
from app.domain.schemas import UserResponse
from app.domain.schemas import VacationResponse
from ..templates import templates_dir
from app.assets.logos import logos_dir
from app.core.config import get_app_settings
from app.core.logging import get_logging
from app.core.celery_worker import celery_app
from app.services import aws

settings = get_app_settings()
log = get_logging(__name__)

def fill_vacations_format(user: User, vacations: VacationResponse):
    
    vacations_dict: dict = vacations.dict()

    userrol = user.userrol[user.active_rol]

    user = UserResponse.from_orm(user).dict(exclude_unset=True)
    
    

    #Rellenar la parte del profesor.
    data_user = {
        "identification": user['identification_number'],
        "full_name": ' '.join([user['names'], user['last_names']]).title(),
        "position": userrol.rol.name,
        #"position": user['rol']['name'],
        "school": user['department']['school']['description'],
        "phone": user['phone'],
        "vinculation_type": user['vinculation_type'],
        "document": user['identification_number'],
        "director_institute_name": user['department']['director'],
        "institute": "Director " + user['department']['description'],
        "dean_school_name": user['department']['school']['dean'],
        "school_position": "Decan@ " + user['department']['school']['description']
    }

    #Rellenar datos de la solicitud.
    today = vacations_dict['created_at']
    initial_date = vacations_dict['vacation']['start_date']
    final_date = vacations_dict['vacation']['end_date']
    data_vacations = {
        "days_type": vacations_dict['application_sub_type']['name'],
        "date_day": str(today.day),
        "date_month": str(today.month),
        "date_year": str(today.year),
        "initial_date_day": str(initial_date.day),
        "initial_date_month": str(initial_date.month),
        "initial_date_year": str(initial_date.year),
        "final_date_day": str(final_date.day),
        "final_date_month": str(final_date.month),
        "final_date_year": str(final_date.year),
        "total_days": str(vacations_dict['vacation']['total_days']),
        "user_signature": vacations_dict['vacation']['signature']
    }

    path = f'user_{user["id"]}/{uuid1()}' + 'formato_vacaciones.xlsx'

    generate_vacations_format_to_aws.apply_async(args=(data_user, data_vacations, path))
    log.debug(path)
    return path

@celery_app.task
def generate_vacations_format_to_aws(user: dict, vacations: dict, path: str, school_id: int = 2):

    cells = open(templates_dir + '/cells_vacation.json')
    cells = json.load(cells)

    wb = load_workbook(filename = templates_dir + '/formato_vacaciones.xlsm')
    target = wb.active

    for datos in [user, vacations]:
        for key in datos.keys():
            if key!='user_signature':
                if cells[key].find(':') < 0:
                    target[cells[key]] = datos[key]
                else:
                    target.merge_cells(cells[key])
                    target[cells[key].split(':')[0]] = datos[key]

    dean_path = logos_dir + f"/signature_{school_id}.png"
    sign_base64 = base64.b64decode(re.sub('data:image/png;base64,',"", vacations['user_signature']))
    img_sign = openpyxl.drawing.image.Image(BytesIO(sign_base64))
    sign_dean = openpyxl.drawing.image.Image(dean_path)
    img_sign.width = 300
    img_sign.height = 320
    target.merge_cells('A21:C24')
    img_sign.anchor = 'A21'
    target.add_image(img_sign)
    target.merge_cells('H23:M23')
    sign_dean.anchor = 'H23'
    target.add_image(sign_dean)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        file = BytesIO(tmp.read())

    aws.s3.push_data_to_s3_bucket(settings.aws_bucket_name, file,
                                  file_name=path, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')