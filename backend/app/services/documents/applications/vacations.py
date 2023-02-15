import json
from datetime import datetime
from uuid import uuid1
from io import BytesIO
from tempfile import NamedTemporaryFile

from openpyxl import load_workbook

from app.domain.models import User, Vacations
from app.domain.schemas import UserResponse
from ..templates import templates_dir
from app.core.config import get_app_settings
from app.core.logging import get_logging
from app.core.celery_worker import celery_app
from app.services import aws


settings = get_app_settings()
log = get_logging(__name__)

def fill_vacations_format(user: User, vacations: Vacations):

    vacations_dict: dict = vacations.dict()

    user = UserResponse.from_orm(user).dict(exclude_unset=True)

    #Rellenar la parte del profesor.
    data_user = {
        "identification": user['identification_number'],
        "full_name": ' '.join([user['names'], user['last_names']]).title(),
        "position": user['rol']['name'],
        "school": user['department']['school']['description'],
        "phone": user['phone'],
        "vinculation_type": user['vinculation_type'],
        "document": user['identification_number'],
        "director_institute_name": user['department']['director'],
        "institute": "Director " + user['department']['description'],
        "dean_school_name": user['department']['school']['dean'],
        "school_position": "Decan@ " + user['department']['school']['description']
    }

    #Rellenar datos de la solicitud.
    today = datetime.now()
    initial_date = datetime.strptime(vacations['initial_date'], '%Y-%m-%dT%H:%M:%S')
    final_date = datetime.strptime(vacations['final_date'], '%Y-%m-%dT%H:%M:%S')
    data_vacations = {
        "days_type": "Días " + vacations['days_type'],
        "date_day": str(today.day),
        "date_month": str(today.month),
        "date_year": str(today.year),
        "initial_date_day": str(initial_date.day),
        "initial_date_month": str(initial_date.month),
        "initial_date_year": str(initial_date.year),
        "final_date_day": str(final_date.day),
        "final_date_month": str(final_date.month),
        "final_date_year": str(final_date.year),
        "total_days": str((final_date - initial_date).days)
    }

    path = f'user_{user["id"]}/{uuid1()}' + 'formato_vacaciones.xlsx'

    generate_vacations_format_to_aws.apply_async(args=(data_user, data_vacations, path))
    log.debug(path)
    return path

@celery_app.task
def generate_vacations_format_to_aws(user: dict, vacations: dict, path: str):

    cells = open(templates_dir + '/cells_vacations.json')
    cells = json.load(cells)

    wb = load_workbook(filename = '/content/formato_vacaciones.xlsm')
    target = wb.active

    for datos in [user, vacations]:
        for key in datos.keys():
            if cells[key].find(':') < 0:
                target[cells[key]] = datos[key]
            else:
                target.merge_cells(cells[key])
                target[cells[key].split(':')[0]] = datos[key]

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        file = BytesIO(tmp.read())

    aws.s3.push_data_to_s3_bucket(settings.aws_bucket_name, file,
                                  file_name=path, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')