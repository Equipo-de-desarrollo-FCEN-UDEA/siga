import json
from datetime import datetime
from uuid import uuid1
from io import BytesIO
from tempfile import NamedTemporaryFile

from openpyxl import load_workbook

from app.domain.models import User, FullTime
from app.domain.schemas import UserResponse
from ..templates import templates_dir
from app.core.config import get_app_settings
from app.core.logging import get_logging
from app.core.celery_worker import celery_app
from app.services import aws


settings = get_app_settings()
log = get_logging(__name__)


def Enumerate_list(lista, index_sep=' ', joint=' '):
    indices = [str(i+1) for i in range(len(lista))]
    result = joint.join([(i+index_sep+content)
                        for i, content in zip(indices, lista)])
    return result


def fill_vice_document(user: User, full_time: FullTime):

    full_time_dict: dict = full_time.dict()

    user = UserResponse.from_orm(user).dict(exclude_unset=True)

    dev_plan = full_time_dict['vice_format']['dev_action_plan']

    action_obj = []
    indic_obj = []
    for obj in dev_plan[0]['objectives']:
        if obj['actions'] and obj['indicators']:
            action_obj.append(obj)
            indic_obj.append(obj)
        elif obj['actions']:
            action_obj.append(obj)
        elif obj['indicators']:
            indic_obj.append(obj)

    data_full_time = {
        # Verificar formato con el modelo de la aplicación. documents/permission date.
        "fecha_diligenciamiento": datetime.now().strftime("%A %d de %B del %Y"),
        "titulo_propuesta": full_time_dict['title'],
        "tiempo_solicitado": str(full_time_dict['vice_format']['time'])+' meses',
        "tema_estrategico": Enumerate_list([jitem['title'] for jitem in dev_plan], index_sep='TE. '),
        "objetivo_estrategico_con_acciones": Enumerate_list([obj['description'] for obj in action_obj], index_sep=').'),
        "objetivo_estrategico_con_indicadores": Enumerate_list([obj['description'] for obj in indic_obj], index_sep=').'),
        "metas": Enumerate_list(list(map(lambda x: x['goal'], full_time_dict['vice_format']['goals']))),
        "acciones_estrategicas": Enumerate_list([action['description'] for obj in action_obj for action in obj['actions']], index_sep=').'),
        "indicadores": Enumerate_list([indic['description'] for obj in indic_obj for indic in obj['indicators']], index_sep='.'),
        "productos": Enumerate_list(list(map(lambda x: x['product'], full_time_dict['vice_format']['products']))),
        "modalidad": full_time_dict['vice_format']['field']
    }
    log.debug(data_full_time)

    data_user = {
        'unidad_academica': '-'.join([user['department']['description'], user['department']['school']['description']]),
        'nombre_completo': ' '.join([user['names'], user['last_names']]).title(),
        '_id': user['identification_number'],
        'num_oficina': user['office'],
        'telefono': user['phone'],
        'email': user['email']
    }

    path = f'user_{user["id"]}/{uuid1()}' + 'formato_vicerrectoria.xlsx'

    generate_vice_format_to_aws.apply_async(args=(data_user, data_full_time, path))
    log.debug(path)
    return path


@celery_app.task
def generate_vice_format_to_aws(user: dict, full_time: dict, path: str):

    cells = open(templates_dir + '/cells_vice.json')
    cells = json.load(cells)
    lone_cells = cells['lone_cells'].copy()
    cells_dic = cells['merge_cells'].copy()
    cells_str = json.dumps(cells_dic)
    cells_str = json.dumps(cells_dic)

    for field, info in zip(full_time.keys(), full_time.values()):
        if info is None:
            continue
        cells_str = cells_str.replace(field, info)

    for field, info in zip(user.keys(), user.values()):
        if info is None:
            continue
        cells_str = cells_str.replace(field, info)

    json_info = json.loads(cells_str)

    wb = load_workbook(filename=templates_dir + '/plantilla_vice.xlsx')
    target = wb.active

    for cells, info in zip(json_info.values(), json_info.keys()):
        target.merge_cells(cells)
        target[cells.split(':')[0]] = info

    for modalidad in lone_cells['modalidad']:
        if modalidad == full_time["modalidad"]:
            target[lone_cells['modalidad'][modalidad]] = 'X'

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        file = BytesIO(tmp.read())

    aws.s3.push_data_to_s3_bucket(settings.aws_bucket_name, file,
                                  file_name=path, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
