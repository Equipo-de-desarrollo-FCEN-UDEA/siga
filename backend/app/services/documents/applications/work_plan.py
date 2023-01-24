import json
from datetime import datetime
from uuid import uuid1
from io import BytesIO
from tempfile import NamedTemporaryFile

from openpyxl import load_workbook

from app.domain.models import User, FullTime
from app.domain.schemas import UserResponse
from ..templates import templates_dir
from app.core.config import get_app_settings
from app.core.logging import get_logging
from app.core.celery_worker import celery_app
from app.services import aws

settings = get_app_settings()
log = get_logging(__name__)


def fill_work_plan_format(user: User, full_time: FullTime) -> str:

    full_time_dict: dict = full_time.dict()
    log.debug(full_time_dict)
    user = UserResponse.from_orm(user).dict(exclude_unset=True)

    data_user = {
        'unidad_academica': user['department']['school']['description'],
        'department': user['department']['description'],
        'cost_center': str(user['department']['cost_center']),
        'nombre_completo': ' '.join([user['last_names'], user['names']]).title(),
        'scale': user['scale'],
        'vinculation_type': user['vinculation_type']
    }

    data_full_time = {
        'period': full_time_dict['work_plan']['period'],
        'date': datetime.now().strftime("%A %d de %B del %Y"),
        'registro': full_time_dict['work_plan']['registro'],
        'partial_time': str(full_time_dict['work_plan']['partial_time']),
        'observations': full_time_dict['work_plan']['observations']
    }

    # Para el tracking.
    tracking_acts = []

    #Identificador de actividades.
    ids_activities = []

    # Total de horas por actividad (Teaching, Research, Extension, Admin, Other).
    total_hours = [0, 0, 0, 0, 0]

    # Sec_2
    teaching_activities = []
    for act in full_time_dict['work_plan']['teaching_activities']:
        if act['level'] == 'pregrado':
            level_pre = 'X'
            level_pos = ''
        elif act['level'] == 'posgrado':
            level_pre = ''
            level_pos = 'X'
        else:
            level_pre = ''
            level_pos = ''

        teaching_activities.append({
            "code_t": act['activity_identification']['code'],
            "group": str(act['activity_identification']['group']),
            "name": act['activity_identification']['name'],
            "student_quantity": str(act['student_quantity']),
            "level_pre": level_pre,
            "level_pos": level_pos,
            "hours_t": str(act['week_hours']['t']),
            "hours_tp": str(act['week_hours']['tp']),
            "hours_p": str(act['week_hours']['p']),
            "total_week_hours": str(act['week_hours']['t'] + act['week_hours']['tp'] + act['week_hours']['p']),
            "total_sem_hours": str(16 * (act['week_hours']['t'] + act['week_hours']['tp'] + act['week_hours']['p']))
        })
        tracking_acts.append(act['activity_tracking'])
        ids_activities.append(act['activity_identification']['code'])
        total_hours[0] += 16*(act['week_hours']['t'] +
                              act['week_hours']['tp']+act['week_hours']['p'])

    # Sec_3
    research_activities = []
    for act in full_time_dict['work_plan']['investigation_activities']:
        research_activities.append({
            "code_i": act['code'],
            "project_identification": act['project_identification'],
            "responsibilities": act['responsibilities'],
            "cost_i": act['cost'],
            "supporting_document": act['supporting_document'],
            "period_hours_i": str(act['period_hours'])
        })
        tracking_acts.append(act['activity_tracking'])
        ids_activities.append(act['code'])
        total_hours[1] += act['period_hours']

    # Sec_4
    extension_activities = []
    for act in full_time_dict['work_plan']['extension_activities']:
        extension_activities.append({
            "code_e": act['code'],
            "activity_identification": act['activity_identification'],
            "responsibility": act['responsibility'],
            "cost_e": act['cost'],
            "week_hours_e": str(act['week_hours']),
            "period_hours_e": str(act['period_hours'])
        })
        tracking_acts.append(act['activity_tracking'])
        ids_activities.append(act['code'])
        total_hours[2] += act['period_hours']

    # Sec_5
    admin_activities = []
    for act in full_time_dict['work_plan']['academic_admin_activities']:
        admin_activities.append({
            "position": act['position'],
            "week_hours_a": str(act['week_hours']),
            "activities": act['activities'],
            "period_hours_a": str(act['period_hours'])
        })
        tracking_acts.append(act['activity_tracking'])
        ids_activities.append("")
        total_hours[3] += act['period_hours']

    # Sec_6
    other_activities = []
    for act in full_time_dict['work_plan']['other_activities']:
        other_activities.append({
            "activity": act['activity'],
            "period_hours_o": act['period_hours'],
        })
        tracking_acts.append(act['activity_tracking'])
        ids_activities.append("")
        total_hours[4] += act['period_hours']

    # Sección 8.
    days = []
    hours_days = []
    for day in full_time_dict['work_plan']['working_week']:
        for part in ['morning', 'afternoon']:
            days.append(day+'_'+part)
            hours_days.append('-'.join([full_time_dict['work_plan']['working_week'][day][part+'_start'],
                                        full_time_dict['work_plan']['working_week'][day][part+'_end']]))

    # Diccionario con todas las actividades.
    data_activities = {'teaching_activities': teaching_activities,
                       'research_activities': research_activities,
                       'extension_activities': extension_activities,
                       'admin_activities': admin_activities,
                       'other_activities': other_activities}

    # Total horas.
    total_hours_semester = dict(zip(['total_hours_t', 'total_hours_r',
                               'total_hours_e', 'total_hours_a', 'total_hours_o'], total_hours))

    # Jornada de trabajo.
    data_day = dict(zip(days, hours_days))

    # Record de actividades.
    record = {'ids_acts':ids_activities, 'act_track': tracking_acts}

    path = f'user_{user["id"]}/{uuid1()}' + 'formato_plan_de_trabajo.xlsx'

    generate_work_plan_format_to_aws.apply_async(
        args=(data_user, data_full_time, data_activities, data_day, total_hours_semester, record, path))

    return path


@celery_app.task
def generate_work_plan_format_to_aws(user: dict, full_time: dict, activities: dict, day: dict, hours_semester: dict, tracking: dict, path: str):

    cells = open(templates_dir + '/cells_plan.json')
    cells = json.load(cells)
    wb = load_workbook(filename=templates_dir + '/plantilla_plan_ext.xlsx')

    target = wb['Cara.1']

    for key in user.keys():
        # Si no se debe hacer merge de las celdas.
        if cells['Cara.1'][key].find(':') < 0:
            target[cells['Cara.1'][key]] = user[key]
        # Si se debe hacer merge.
        else:
            target.merge_cells(cells['Cara.1'][key])
            target[cells['Cara.1'][key].split(':')[0]] = user[key]

    for key in list(full_time.keys())[:-1]:
        # Si no se debe hacer merge de las celdas.
        if cells['Cara.1'][key].find(':') < 0:
            target[cells['Cara.1'][key]] = full_time[key]
        # Si se debe hacer merge.
        else:
            target.merge_cells(cells['Cara.1'][key])
            target[cells['Cara.1'][key].split(':')[0]] = full_time[key]

    for activity in list(activities.keys())[:3]:
        index = cells['Cara.1']['multiple_rows'][activity]['range_cells']
        content = activities[activity]
        for i in range(len(content)):
            for key in content[i].keys():
                coord = cells['Cara.1']['multiple_rows'][activity][key]
                if coord.find(':') < 0:
                    coord = coord+str(index+i)
                    target[coord] = content[i][key]
                else:
                    coord = ':'.join([l+str(index+i) for l in coord.split(':')])
                    target.merge_cells(coord)
                    target[coord.split(':')[0]] = content[i][key]

    for key in list(hours_semester.keys())[:3]:
        # Si no se debe hacer merge de las celdas.
        if cells['Cara.1']['total_hours'][key].find(':') < 0:
            target[cells['Cara.1']['total_hours'][key]] = hours_semester[key]
        # Si se debe hacer merge.
        else:
            target.merge_cells(cells['Cara.1']['total_hours'][key])
            target[cells['Cara.1']['total_hours'][key].split(':')[0]] = hours_semester[key]

    target = wb['Cara.2']

    for activity in list(activities.keys())[-2:]:
        index = cells['Cara.2']['multiple_rows'][activity]['range_cells']
        content = activities[activity]
        for i in range(len(content)):
            for key in content[i].keys():
                coord = cells['Cara.2']['multiple_rows'][activity][key]
            if coord.find(':') < 0:
                coord = coord+str(index+i)
                target[coord] = content[i][key]
            else:
                coord = ':'.join([l+str(index+i) for l in coord.split(':')])
                target.merge_cells(coord)
                target[coord.split(':')[0]] = content[i][key]

    index = cells['Cara.2']['multiple_rows']['activity_tracking']['range_cells']
    identifications = tracking[list(tracking.keys())[0]]
    activities_track = tracking[list(tracking.keys())[1]]
    for i in range(len(activities_track)):
        for key in activities_track[i].keys():
            if key in ['date_1', 'date_2']:
                continue
            coord = cells['Cara.2']['multiple_rows']['activity_tracking'][key]
            coord_idx = cells['Cara.2']['multiple_rows']['activity_tracking']['id_activity']

            coord_idx = ':'.join([l+str(index+i) for l in coord_idx.split(':')])
            target.merge_cells(coord_idx)
            target[coord_idx.split(':')[0]] = identifications[i]

            if coord.find(':') < 0:
                coord = coord+str(index+i)
                target[coord] = activities_track[i][key]
            else:
                coord = ':'.join([l+str(index+i) for l in coord.split(':')])
                target.merge_cells(coord)
                target[coord.split(':')[0]] = activities_track[i][key]

    for key in list(hours_semester.keys())[-2:]:
        if cells['Cara.2']['total_hours'][key].find(':') < 0:
            target[cells['Cara.2']['total_hours'][key]] = hours_semester[key]
        else:
            target.merge_cells(cells['Cara.2']['total_hours'][key])
            target[cells['Cara.2']['total_hours'][key].split(':')[0]] = hours_semester[key]

    for daypart, rango in day.items():
        target[cells['Cara.2'][daypart]] = rango
    
    target.merge_cells(cells['Cara.2']['observations'])
    target[cells['Cara.2']['observations'].split(':')[0]] = full_time['observations']

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        file = BytesIO(tmp.read())

    aws.s3.push_data_to_s3_bucket(settings.aws_bucket_name, file,
                                  file_name=path, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
